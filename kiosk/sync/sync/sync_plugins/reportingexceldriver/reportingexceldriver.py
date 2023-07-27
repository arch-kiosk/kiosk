import logging
import os
import shutil
from copy import copy
from typing import Union, Callable, List, Iterator, Tuple

from flake8.api.legacy import Report
from openpyxl.worksheet.worksheet import Worksheet

import kioskstdlib
from reportingdock.reportinglib import ReportingException
from sync_plugins.reportingexceldriver.reportingmapperexcel import ReportingMapperExcel
from sync_plugins.reportingdock.reportingoutputdriver import ReportingOutputDriver
from synchronization import Synchronization
from synchronizationplugin import SynchronizationPlugin
from openpyxl import load_workbook, Workbook


# ************************************************************************
# Plugin code for ReportingExcelDriver
# ************************************************************************
class PluginReportingExcelDriver(SynchronizationPlugin):
    _plugin_version = 0.1

    def all_plugins_ready(self):
        app: Synchronization = self.app
        if app:
            ReportingExcelDriver.register(app.type_repository)
            logging.debug("PluginReportingExcelDriver: plugin and driver type registered")
        else:
            logging.error("PluginReportingExcelDriver: plugin and driver type could not be registered due to no app.")
            return False

        return True


# ************************************************************************
# ReportingExcelDriver
# ************************************************************************

class ExcelReference:
    TYPE_CELL = 0
    TYPE_RANGE = 1

    def __init__(self, ref_str: str):
        ref1 = ref_str.split("!")
        self.type = self.TYPE_CELL
        self.sheet = ""
        self.cell_str = []

        if len(ref1) == 1:
            ref_str = ref1[0]
        else:
            self.sheet = ref1[0]
            ref_str = ref1[1]

        for ref in ref_str.split(":"):
            self.cell_str.append(ref)

        if len(self.cell_str) > 1:
            self.type = self.TYPE_RANGE


class ReportingExcelDriver(ReportingOutputDriver):
    ANNOT_KEY = '/Annots'
    ANNOT_FIELD_KEY = '/T'
    ANNOT_VAL_KEY = '/V'
    ANNOT_RECT_KEY = '/Rect'
    SUBTYPE_KEY = '/Subtype'
    WIDGET_SUBTYPE_KEY = '/Widget'

    can_view = False
    can_download = True
    can_zip = True

    @classmethod
    def get_supported_file_extensions(cls):
        return ["xlsx"]

    def _get_mapper_class(self):
        return ReportingMapperExcel

    def _prepare_file(self):
        if not os.path.exists(self.target_dir):
            os.mkdir(self.target_dir)
        target_filename = os.path.join(self.target_dir, self.target_file_name_without_extension + ".xlsx")
        shutil.copyfile(self.template_file, target_filename)
        logging.info(f"ReportingExcelDriver._prepare_file: creating new workbook {self.template_file}")
        return target_filename

    def _map_to_cell(self, wb: Workbook, ref: ExcelReference, value):
        cell_address = ""
        if ref.type == ref.TYPE_RANGE:
            raise ReportingException(f"{self.__class__.__name__}.map_to_cell: the addressed name "
                                     f"'{ref.cell_str}' is a range in Excel. Single values cannot be "
                                     f"mapped onto ranges.")
        try:
            dests = wb.defined_names[ref.cell_str[0]].destinations
            c = 0
            for sheet_title, coord in dests:
                c += 1
                if c > 1:
                    raise ReportingException(f"{self.__class__.__name__}.map_to_cell: the addressed name "
                                             f"'{ref.cell_str}' is a range in Excel. Single values cannot be "
                                             f"mapped onto ranges.")
                ws = wb[sheet_title]
                cell_address = ws[coord].coordinate
                ref.sheet = sheet_title
        except KeyError:
            # not a defined name in workbook but also not in range notation
            cell_address = ref.cell_str[0]

        except BaseException as e:
            raise ReportingException(f"{self.__class__.__name__}.map_to_cell: "
                                     f"{repr(e)} when trying name '{ref.cell_str[0]}'")

        if ref.sheet:
            ws = wb[ref.sheet]
        else:
            ws = wb.active

        ws[cell_address] = value

    def _copy_range_with_format(self, ws: Worksheet, source_range: ExcelReference, dest_range: ExcelReference):
        if source_range.type != source_range.TYPE_RANGE:
            raise ReportingException(f"{self.__class__.__name__}._copy_range_with_format: "
                                     f"source_range is not a range.")

        src = ws[source_range.cell_str[0]:source_range.cell_str[1]]
        dst = ws[dest_range.cell_str[0]:dest_range.cell_str[1]]
        dst_row = dst[0][0].row
        for src_row in src:
            dst_col = dst[0][0].col_idx
            for cell in src_row:
                new_cell = ws.cell(dst_row, dst_col)
                new_cell.value = cell.value
                new_cell._style = copy(cell._style)
                dst_col += 1
            dst_row += 1

    def _add_row_below(self, ws: Worksheet, source_ref: ExcelReference):
        row_number = ws[source_ref.cell_str[0]].row
        ws.insert_rows(row_number + 1)

    def _get_table_name_if_exists(self, value):
        """
        checks whether the value refers to an existing table in the mapping definition and returns
        the table (without the # prefix)
        :param value: a table name in # notation
        :return: the table name without # or an empty string if no table was found
        """
        if value and value[0] == '#':
            table_name = value[1:]
            if "tables" in self.mapping_definition and table_name in self.mapping_definition["tables"]:
                return table_name

        return ""

    def _map_table(self, wb: Workbook, ref: ExcelReference,
                   table_name: str,
                   _on_load_records: Union[Callable[[str, List[str], List[str]], Iterator[Tuple]], None],
                   add_rows=True, call_count=0):
        table_def = self.mapping_definition["tables"][table_name]
        if call_count > 10:
            raise ReportingException(f"{self.__class__.__name__}._map_table: Endless recursion detected.")

        cell_address = ""
        if ref.type == ref.TYPE_RANGE:
            raise ReportingException(f"{self.__class__.__name__}._map_table_to_range: the addressed name "
                                     f"'{ref.cell_str}' is a range in Excel. Please state the upper left cell "
                                     f"for the table instead.")
        try:
            dests = wb.defined_names[ref.cell_str[0]].destinations
            c = 0
            for sheet_title, coord in dests:
                c += 1
                if c > 1:
                    raise ReportingException(f"{self.__class__.__name__}.map_to_cell: the addressed name "
                                             f"'{ref.cell_str}' is a range in Excel. Please state the upper left cell "
                                             f"for the table instead.")
                ws = wb[sheet_title]
                cell_address = ws[coord].coordinate
                ref.sheet = sheet_title
        except KeyError:
            # not a defined name in workbook but also not in range notation
            cell_address = ref.cell_str[0]

        except BaseException as e:
            raise ReportingException(f"{self.__class__.__name__}.map_to_cell: "
                                     f"{repr(e)} when trying name '{ref.cell_str[0]}'")

        if ref.sheet:
            ws = wb[ref.sheet]
        else:
            ws = wb.active

        cells = table_def["cells"]
        mapping = table_def["mapping"]
        current_cell = ws[cell_address]
        xl_top_row = current_cell.row
        xl_left_col = current_cell.col_idx
        xl_right_col = xl_left_col + len(cells)
        xl_current_row = xl_top_row
        xl_current_col = xl_left_col
        top_range = cell_address + ":" + ws.cell(row=xl_top_row, column=xl_right_col).coordinate
        templated = add_rows and (kioskstdlib.to_bool(table_def["templated"]) if "templated" in table_def else False)

        if "sources" in table_def:
            tables = table_def["sources"]
        else:
            tables = [table_name]
        if "order_by" in table_def:
            order_by = table_def["order_by"]
        else:
            order_by = []

        if "row_offset" in table_def:
            xl_current_row += int(table_def["row_offset"])

        for table in tables:
            for row in _on_load_records(table, [], order_by):
                if add_rows:
                    self._add_row_below(ws, ExcelReference(ws.cell(xl_current_row, xl_current_col).coordinate))
                    xl_current_row += 1
                    if templated:
                        dest_range = f"{ws.cell(row=xl_current_row, column=xl_left_col).coordinate}:" \
                                     f"{ws.cell(row=xl_current_row, column=xl_right_col).coordinate}"
                        self._copy_range_with_format(ws,
                                                     source_range=ExcelReference(top_range),
                                                     dest_range=ExcelReference(dest_range))
                for cell_idx, cell in enumerate(cells):
                    if cell in mapping:
                        cell_value = self._mapper.get_table_cell_mapping(mapping[cell], row)
                        if cell_value is None:
                            cell_value = ''
                        ws.cell(row=xl_current_row, column=xl_left_col + cell_idx, value=cell_value)

                if not add_rows:
                    xl_current_row += 1

        if "append_table" in table_def:
            next_table = table_def["append_table"]
            if next_table and next_table in self.mapping_definition["tables"]:
                table_origin = ExcelReference(ws.title + "!" + ws.cell(row=xl_current_row+1, column=xl_current_col).coordinate)
                self._map_table(wb, table_origin, next_table, _on_load_records=_on_load_records, add_rows=False,
                                call_count=call_count + 1)
            else:
                raise ReportingException(f"{self.__class__.__name__}._map_table: "
                                         f"append_table refers to missing table '{next_table}'.")
        if templated:
            ws.delete_rows(xl_top_row)

    def _get_table_cell_mapping(self, row, cell):
        return cell

    def execute(self, _on_load_records: Union[Callable[[str, List[str]], Iterator[Tuple]], None]) -> str:
        """
        executes the report
        :param _on_load_records: a callable which gets the table_name and the "columns" list and
                                 returns a generator that returns a new row (a tuple with values)
                                 on each iteration
        :returns the path and filename of the target file
        """

        self.check_premisses()
        if not self._mapper:
            raise ReportingException(f"{self.__class__.__name__}.execute: "
                                     f"execute can only work if map has been used before.")
        target_filename = self._prepare_file()
        wb = load_workbook(target_filename)
        sheet_titles = {}
        for key, value in self._mapped_values.items():
            if key.strip()[-1:] == '!':
                # a sheet reference
                sheet_name = key.strip()[:-1]
                sheet_titles[sheet_name] = value
            else:
                ref = ExcelReference(key)
                table_name = self._get_table_name_if_exists(value)
                if table_name:
                    self._map_table(wb, ref, table_name, _on_load_records)
                else:
                    self._map_to_cell(wb, ref, value)

        self._change_sheet_titles(wb, sheet_titles)
        wb.save(target_filename)
        return target_filename

    def _change_sheet_titles(self, wb: Workbook, sheet_titles: dict):
        try:
            for old_title, new_title in sheet_titles.items():
                ws = wb[old_title]
                ws.title = new_title
        except BaseException as e:
            logging.warning(f"{self.__class__.__name__}._change_sheet_titles: {repr(e)}")
